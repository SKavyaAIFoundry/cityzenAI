import os
import shutil

from openpyxl import load_workbook  # MIT/Expat licence
from zipfile import ZipFile


# methods for extracting content from Excel files

def analyseXlsxFile(fileName):
    """
    Function for extracting content from Excel documents.
    Returns the following:
    - cell data as text content
    - word count
    - row count
    - unique image count
    """

    word_count = 0
    paragraph_count = 0     # returns total number of rows in all sheets
    bullet_count = 0        # not yet implemented
    image_count = 0

    # data_only=False to extract formulas, True to extract results 
    workbook = load_workbook(fileName, data_only=True)

    document_text = ''

    # extract data from each worksheet
    for worksheet in workbook.worksheets:
        # get number of rows and columns in worksheet
        worksheet_rows = worksheet.max_row
        worksheet_columns = worksheet.max_column
        # for each row...
        for i in range(1, worksheet_rows + 1):
            paragraph_count += 1 # based on total row count in all worksheets
            # ... extract each column value
            for j in range(1, worksheet_columns + 1):
                if worksheet.cell(row=i, column=j).value != None:
                    document_text += str(worksheet.cell(row=i, column=j).value) + " "

    document_words = document_text.split()

    word_count = len(document_words)

    # get number of images by counting extracted image files...

    # path for temp excel content directory
    directory_path = './temp_excel_content'
    # path for directory containing image files
    media_path = directory_path + '/xl/media'

    # extract file contents to temp directory
    with ZipFile(fileName, 'r') as zip_file:
        zip_file.extractall(directory_path)

    # image file extensions to search for
    image_file_extensions = ['.gif', '.jpg', '.png']

    # search media directory for image files
    if os.path.exists(media_path):
        for file_name in os.listdir(media_path):
            file_path = os.path.join(media_path, file_name)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    file_full_name, file_extension = os.path.splitext(file_name)
                    if file_extension in image_file_extensions:
                        image_count += 1
            except Exception as exception:
                print('Error with file: %s.  Exception: %s' % (file_path, exception))
    # delete temp directory when finished
    shutil.rmtree(directory_path)

    return document_text, word_count, paragraph_count, bullet_count, image_count


def xlsxRowsToTextFile(fileName):
    """
    Function for extracting content from Excel workbooks to text files
    Creates a text file for each sheet in the workbook
    Creates a line of text for each row of the individual worksheets
    Separates column cell content with commas
    Empty cells are stored to retain column data structure
    """

    # data_only=False to extract formulas, True to extract results 
    workbook = load_workbook(fileName, data_only=True)

    file_full_name, file_extension = os.path.splitext(fileName)

    for worksheet in workbook.worksheets:
        # get number of rows and columns in worksheet
        worksheet_rows = worksheet.max_row
        worksheet_columns = worksheet.max_column

        worksheet_name = worksheet.title
        # 'w+' to overwrite files, 'a' to append to files
        outputFile = open('{}_sheet_{}.txt'.format(file_full_name, worksheet_name), 'w+')

        # extract row content to line of text file
        for i in range(1, worksheet_rows + 1):
            for j in range(1, worksheet_columns + 1):
                if worksheet.cell(row=i, column=j).value != None:
                    print(worksheet.cell(row=i, column=j).value)
                    outputFile.write(str(worksheet.cell(row=i, column=j).value) + ",")
                else:
                    outputFile.write(",")
            outputFile.write("\n")
        outputFile.close()


